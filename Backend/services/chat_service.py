import os
import re
from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Any, Dict, List, Optional

import requests
from openpyxl import Workbook, load_workbook
from requests import RequestException

OPENAI_CHAT_COMPLETIONS_URL = "https://api.openai.com/v1/chat/completions"
ALLOWED_HISTORY_ROLES = {"user", "assistant"}
CHAT_LOG_FILE_NAME = "AB_Deliveries_Chatbot_Logs.xlsx"
CHAT_LOG_HEADERS = ["Timestamp", "Caller Name", "Phone Number", "Conversation Details"]
PHONE_PATTERNS = [
    re.compile(r"(?<!\d)(?:\+972[-\s]?|0)?5\d[-\s]?\d{7}(?!\d)"),
    re.compile(r"(?<!\d)\d{7,15}(?!\d)"),
]
_LOG_WRITE_LOCK = Lock()


def _normalize_text(raw_value: Any) -> str:
    return str(raw_value or "").strip()


def _resolve_chat_log_path() -> Path:
    project_root = Path(__file__).resolve().parents[2]
    return project_root / "ToastServer" / CHAT_LOG_FILE_NAME


def _normalize_history(raw_history: Any) -> List[Dict[str, str]]:
    if not isinstance(raw_history, list):
        return []

    normalized: List[Dict[str, str]] = []
    for item in raw_history[-12:]:
        if not isinstance(item, dict):
            continue

        role = _normalize_text(item.get("role")).lower()
        content = _normalize_text(item.get("content"))
        if role not in ALLOWED_HISTORY_ROLES or not content:
            continue

        normalized.append({"role": role, "content": content})

    return normalized


def _normalize_phone(raw_value: str) -> str:
    digits = "".join(character for character in raw_value if character.isdigit() or character == "+")
    if digits.startswith("972") and not digits.startswith("+972"):
        digits = f"+{digits}"
    return digits


def _extract_phone_number(user_payload: Optional[Dict[str, Any]], conversation_text: str) -> str:
    if isinstance(user_payload, dict):
        for field_name in ("phone", "phone_number", "phoneNumber", "mobile"):
            candidate = _normalize_text(user_payload.get(field_name))
            if not candidate:
                continue
            for pattern in PHONE_PATTERNS:
                match = pattern.search(candidate)
                if match:
                    return _normalize_phone(match.group(0))

    for pattern in PHONE_PATTERNS:
        match = pattern.search(conversation_text)
        if match:
            return _normalize_phone(match.group(0))

    return ""


def _build_caller_name(user_payload: Optional[Dict[str, Any]]) -> str:
    if not isinstance(user_payload, dict):
        return "Unknown"

    first_name = _normalize_text(
        user_payload.get("first_name") or user_payload.get("firstName")
    )
    last_name = _normalize_text(
        user_payload.get("last_name") or user_payload.get("lastName")
    )
    full_name = " ".join(part for part in [first_name, last_name] if part).strip()
    return full_name or "Unknown"


def _build_conversation_details(
    history: List[Dict[str, str]],
    user_message: str,
    assistant_message: str = "",
    error_message: str = "",
) -> str:
    lines: List[str] = []
    for item in history[-6:]:
        role_label = "Caller" if item["role"] == "user" else "Assistant"
        lines.append(f"{role_label}: {item['content']}")

    lines.append(f"Caller: {user_message}")

    if assistant_message:
        lines.append(f"Assistant: {assistant_message}")

    if error_message:
        lines.append(f"Error: {error_message}")

    details = "\n".join(lines).strip()
    return details[:8000]


def _append_chat_log_row(caller_name: str, phone_number: str, conversation_details: str) -> None:
    log_path = _resolve_chat_log_path()
    log_path.parent.mkdir(parents=True, exist_ok=True)

    with _LOG_WRITE_LOCK:
        created_new_workbook = False
        if log_path.exists():
            workbook = load_workbook(log_path)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Logs"
            created_new_workbook = True

        first_row_values = [worksheet.cell(row=1, column=index).value for index in range(1, 5)]
        if created_new_workbook or (worksheet.max_row == 1 and all(value is None for value in first_row_values)):
            for column_index, header in enumerate(CHAT_LOG_HEADERS, start=1):
                worksheet.cell(row=1, column=column_index, value=header)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_values = [timestamp, caller_name, phone_number, conversation_details]
        if worksheet.max_column > len(row_values):
            row_values.extend([""] * (worksheet.max_column - len(row_values)))

        worksheet.append(row_values)
        workbook.save(log_path)
        workbook.close()


def _log_chat_conversation(
    payload: Dict[str, Any],
    history: List[Dict[str, str]],
    user_message: str,
    assistant_message: str = "",
    error_message: str = "",
) -> None:
    try:
        user_payload = payload.get("user") if isinstance(payload, dict) else {}
        details = _build_conversation_details(history, user_message, assistant_message, error_message)
        caller_name = _build_caller_name(user_payload)
        phone_number = _extract_phone_number(user_payload, details)
        _append_chat_log_row(caller_name, phone_number, details)
    except Exception:
        # Logging failure must not block chat flow.
        pass


def _build_system_prompt(user_payload: Optional[Dict[str, Any]]) -> str:
    first_name = _normalize_text((user_payload or {}).get("first_name"))

    base_prompt = (
        "אתה נציג שירות ומכירות של A.B Deliveries.\n"
        "חובות התפקיד שלך:\n"
        "1) שירות לקוחות בנושא סטטוס משלוחים וחבילות.\n"
        "2) תמיכה מכירתית שמעודדת את הלקוח להזמין יותר משלוחים בצורה נעימה ולא אגרסיבית.\n\n"
        "כללי שפה והצגה:\n"
        "- כתוב בעברית בלבד.\n"
        "- גם אם המשתמש כותב באנגלית או שפה אחרת, השב בעברית בלבד.\n"
        "- שמור על ניסוח ברור, ידידותי ומקצועי.\n"
        "- השתמש בפורמט שמתאים ל-RTL (משפטים קצרים, רשימות קצרות כשצריך).\n\n"
        "כללי שירות:\n"
        "- כשמבקשים סטטוס חבילה, בקש פרטים מזהים חסרים בצורה ממוקדת (למשל מספר מעקב/מספר הזמנה/טלפון).\n"
        "- אם אין מספיק מידע, הסבר מה חסר ובקש את המינימום הנדרש להמשך.\n"
        "- אל תמציא סטטוס משלוח שלא נמסר במפורש.\n\n"
        "כללי מכירה:\n"
        "- בכל תשובה נסה להוסיף הצעה קצרה ורלוונטית להזמנה נוספת או לשדרוג שירות משלוחים.\n"
        "- הדגש ערך עסקי: חיסכון בזמן, אמינות, איסוף מהיר, ושירות לעסקים.\n"
    )

    if first_name:
        return (
            f"{base_prompt}\n"
            f"שם המשתמש הוא {first_name}. פנה אליו בשמו הפרטי בצורה טבעית ונעימה."
        )

    return base_prompt


def generate_chat_reply(payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not payload:
        return {"success": False, "status_code": 400, "error": "No data sent"}

    message = _normalize_text(payload.get("message"))
    if not message:
        return {"success": False, "status_code": 400, "error": "Message is required"}

    history = _normalize_history(payload.get("history"))
    api_key = _normalize_text(os.getenv("OPENAI_API_KEY"))
    if not api_key:
        error_message = "OpenAI is not configured on the server."
        _log_chat_conversation(payload, history, message, error_message=error_message)
        return {
            "success": False,
            "status_code": 503,
            "error": error_message,
        }

    model = _normalize_text(os.getenv("OPENAI_MODEL")) or "gpt-4o-mini"
    system_prompt = _build_system_prompt(payload.get("user"))

    messages = [{"role": "system", "content": system_prompt}, *history, {"role": "user", "content": message}]

    try:
        response = requests.post(
            OPENAI_CHAT_COMPLETIONS_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": model,
                "messages": messages,
                "temperature": 0.6,
            },
            timeout=30,
        )
    except RequestException:
        error_message = "Failed to reach OpenAI service."
        _log_chat_conversation(payload, history, message, error_message=error_message)
        return {
            "success": False,
            "status_code": 502,
            "error": error_message,
        }

    if not response.ok:
        try:
            error_payload = response.json()
            error_message = _normalize_text((error_payload.get("error") or {}).get("message"))
        except ValueError:
            error_message = ""

        error_text = error_message or "OpenAI request failed."
        _log_chat_conversation(payload, history, message, error_message=error_text)
        return {
            "success": False,
            "status_code": 502,
            "error": error_text,
        }

    try:
        payload_json = response.json()
    except ValueError:
        error_message = "Invalid response from OpenAI service."
        _log_chat_conversation(payload, history, message, error_message=error_message)
        return {
            "success": False,
            "status_code": 502,
            "error": error_message,
        }

    choices = payload_json.get("choices")
    if not isinstance(choices, list) or len(choices) == 0:
        error_message = "OpenAI returned no response choices."
        _log_chat_conversation(payload, history, message, error_message=error_message)
        return {
            "success": False,
            "status_code": 502,
            "error": error_message,
        }

    assistant_message = _normalize_text(((choices[0] or {}).get("message") or {}).get("content"))
    if not assistant_message:
        error_message = "OpenAI returned an empty response."
        _log_chat_conversation(payload, history, message, error_message=error_message)
        return {
            "success": False,
            "status_code": 502,
            "error": error_message,
        }

    _log_chat_conversation(payload, history, message, assistant_message=assistant_message)
    return {
        "success": True,
        "status_code": 200,
        "data": {"reply": assistant_message},
    }
